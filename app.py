# from flask import Flask, render_template, request, send_file, redirect, url_for, flash
# import pandas as pd
# import tempfile
# import os
# import re
# from openpyxl import load_workbook
# from openpyxl.styles import Protection

# app = Flask(__name__)
# app.secret_key = "super_secret_key_please_change"

# def normalize_key(s):
#     return re.sub(r"[\s_]+", "", s).lower()

# # Required columns and fuzzy variants
# required_columns_map = {
#     "TEST ID": ["testid", "test id", "test_id", "id"],
#     "TIME STAMP": ["timestamp", "time stamp", "time_stamp"],
#     "COARSE TEST": ["coarsetest", "coarse test"],
#     "PRESSURE TEST": ["pressuretest", "pressure test"],
#     "BACKGROUND TEST": ["backgroundtest", "background test"],
#     "LEAK TEST(COARSE)": ["leaktest(coarse)", "leak test coarse"],
#     "LEAK TEST(FINE)": ["leaktest(fine)", "leak test(fine)", "leak test fine", "leak_test_fine", "leaktestfine", "leak test-fine", "leak (fine) test", "fineleaktest", "fine leak test"],
#     "TEST RESULT": ["testresult", "test result", "result"],
#     "COARSE TEST START PRESSURE": ["coarsestartpressure", "coarse test start pressure"],
#     "COARSE TEST END PRESSURE": ["coarseendpressure", "coarse test end pressure"],
#     "COARSE TEST PRESSURE DROP": ["coarsepressuredrop", "coarse test pressure drop"],
#     "PRESSURE TEST START PRESSURE": ["pressurestartpressure", "pressure test start pressure"],
#     "PRESSURE TEST END PRESSURE": ["pressureendpressure", "pressure test end pressure"],
#     "PRESSURE TEST PRESSURE DROP": ["pressurepressuredrop", "pressure test pressure drop"],
#     "HE BACKGROUND": ["hebackground", "he background"],
#     "ACTIVE MACHINE FACTOR": ["activemachinefactor", "active machine factor"],
#     "HELIUM CONCENTERATION": ["heliumconcenteration", "helium concentration"],
#     "HE COARSE FILL PRESSURE": ["hecoarsefillpressure", "he coarse fill pressure"],
#     "HE COARSE TEST ELEPSED TIME": ["hecoarsetestelepsedtime", "he coarse test elepsed time"],
#     "COARSE LEAK RATE(RAW VALUE)": ["coarseleakrate(rawvalue)", "coarse leak rate raw"],
#     "COARSE LEAK RATE(CORRECTED)": ["coarseleakrate(corrected)", "coarse leak rate corrected"],
#     "COARSE LEAK TEST VACUUM": ["coarseleaktestvacuum", "coarse leak test vacuum", "coarse_leak_test_vacuum", "coarseleaktest vaccum", "coarse leak test vaccum", "coarse leak test vaccuum", "coarse test vacuum", "coarsevacuumtest", "coarse test vaccum"],
#     "HE FINE FILL PRESSURE": ["hefinefillpressure", "he fine fill pressure"],
#     "HE FINE TEST ELEPSED TIME": ["hefinetestelepsedtime", "he fine test elepsed time"],
#     "FINE LEAK RATE(RAW VALUE)": ["fineleakrate(rawvalue)", "fine leak rate raw"],
#     "FINE LEAK RATE(CORRECTED)": ["fineleakrate(corrected)", "fine leak rate corrected"],
#     "FINE LEAK TEST VACUUM": ["fineleaktestvacuum", "fine leak test vacuum", "fine_leak_test_vacuum", "fine leak testvaccum", "fineleaktest vaccum", "fineleaktest vaccuum", "fine leak test vaccuum", "fine test vacuum", "finevacuumtest", "fine test vaccum"],
#     "DRAWING NUMBER": ["drawingnumber", "drawing number"],
#     "REFERENCE DOCUMENT": ["referencedocument", "reference document"],
#     "CUSTOMER NAME": ["customername", "customer name"],
#     "RECIPE NAME": ["recipename", "recipe name"],
#     "COMPONENT NAME": ["componentname", "component name"],
#     "SUPERVISOR": ["supervisor", "manager"]
# }

# editable_columns = ["WORK ORDER.", "HOUSING NO.", "FLANG NO.", "PROJECT", "TYPE", "REMARKS"]

# # Lookup map
# normalized_required_lookup = {}
# for output_col, variants in required_columns_map.items():
#     for variant in variants:
#         normalized_required_lookup[normalize_key(variant)] = output_col

# def parse_test_ids(input_str):
#     return sorted(set(p.strip() for p in input_str.split(',') if p.strip()))

# @app.route("/", methods=["GET", "POST"])
# def index():
#     if request.method == "POST":
#         test_ids_raw = request.form.get("emp_id")
#         file = request.files.get("file")

#         if not file or file.filename == "":
#             flash("❌ Please upload a valid Excel file.")
#             return redirect(url_for("index"))

#         try:
#             with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp:
#                 file.save(temp.name)
#                 df_raw = pd.read_excel(temp.name, dtype=str)

#             col_mapping = {}
#             for col in df_raw.columns:
#                 norm = normalize_key(col)
#                 if norm in normalized_required_lookup:
#                     col_mapping[col] = normalized_required_lookup[norm]

#             df = df_raw.rename(columns=col_mapping)

#             for standard_col in required_columns_map.keys():
#                 if standard_col not in df.columns:
#                     df[standard_col] = pd.NA

#             if "TEST ID" not in df.columns:
#                 flash(f"❌ 'TEST ID' column not found. Found: {list(df.columns)}")
#                 return redirect(url_for("index"))

#             df["TEST ID"] = df["TEST ID"].astype(str).str.strip()

#             test_ids = parse_test_ids(test_ids_raw)
#             available_ids = set(df["TEST ID"].unique())
#             requested_ids = set(test_ids)
#             invalid_ids = requested_ids - available_ids

#             if invalid_ids:
#                 flash(f"❌ These Test ID(s) were not found: {', '.join(invalid_ids)}")
#                 return redirect(url_for("index"))

#             filtered = df[df["TEST ID"].isin(test_ids)]
#             ordered_cols = ["TEST ID"] + [col for col in required_columns_map if col != "TEST ID"]
#             final_df = filtered[ordered_cols]
#             final_df.insert(0, "S.No", range(1, len(final_df) + 1))

#             for col in editable_columns:
#                 final_df[col] = ""

#             output_file = os.path.join(tempfile.gettempdir(), "testid_export.xlsx")
#             final_df.to_excel(output_file, index=False)

#             wb = load_workbook(output_file)
#             ws = wb.active

#             editable_col_indices = [
#                 idx + 1 for idx, col in enumerate(final_df.columns)
#                 if col in editable_columns or col == "S.No"
#             ]

#             for row in ws.iter_rows():
#                 for cell in row:
#                     if cell.row == 1 or cell.column in editable_col_indices:
#                         cell.protection = Protection(locked=False)
#                     else:
#                         cell.protection = Protection(locked=True)

#             ws.protection.sheet = True
#             ws.protection.password = "locked123"

#             wb.save(output_file)

#             filename = f"testdata_{'_'.join(test_ids)}.xlsx"
#             return send_file(output_file, as_attachment=True, download_name=filename)

#         except Exception as e:
#             flash(f"❌ Error processing file: {str(e)}")
#             return redirect(url_for("index"))

#     return render_template("index.html")

# if __name__ == "__main__":
#     app.run(debug=True)
from flask import Flask, render_template, request, send_file, flash, redirect, url_for
import os
import tempfile
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Protection
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as ExcelImage
import re

app = Flask(__name__)
app.secret_key = "super_secret_key_please_change"

def normalize_key(s):
    return re.sub(r"[\s_\-]+", "", s).lower()

cell_map = {
    "IDENTIFICATION NUMBER": "C", "TIME STAMP": "H", "COARSE TEST": "J", "PRESSURE TEST": "K",
    "BACKGROUND TEST": "L", "LEAK TEST(COARSE)": "M", "LEAK TEST(FINE)": "N", "TEST RESULT": "O",
    "COARSE TEST START PRESSURE": "P", "COARSE TEST END PRESSURE": "Q", "COARSE TEST PRESSURE DROP": "R",
    "PRESSURE TEST START PRESSURE": "S", "PRESSURE TEST END PRESSURE": "T", "PRESSURE TEST PRESSURE DROP": "U",
    "HE BACKGROUND": "V", "ACTIVE MACHINE FACTOR": "W", "HELIUM CONCENTERATION": "X",
    "HE COARSE FILL PRESSURE": "Y", "HE COARSE TEST ELEPSED TIME": "Z", "COARSE LEAK RATE(RAW VALUE)": "AA",
    "COARSE LEAK RATE(CORRECTED)": "AB", "COARSE LEAK TEST VACUUM": "AC", "HE FINE FILL PRESSURE": "AD",
    "HE FINE TEST ELEPSED TIME": "AE", "FINE LEAK RATE(RAW VALUE)": "AF", "FINE LEAK RATE(CORRECTED)": "AG",
    "FINE LEAK TEST VACUUM": "AH", "DRAWING NUMBER": "AI", "REFERENCE DOCUMENT": "AJ",
    "CUSTOMER NAME": "AK", "RECIPE NAME": "AL", "COMPONENT NAME": "AM", "SUPERVISOR": "AN"
}

column_variants = {
    "IDENTIFICATION NUMBER": ["identificationnumber", "identification number", "id number", "idno", "id_no", "ident_no"],
    "TIME STAMP": ["timestamp", "time stamp", "time_stamp"],
    "COARSE TEST": ["coarsetest", "coarse test"],
    "PRESSURE TEST": ["pressuretest", "pressure test"],
    "BACKGROUND TEST": ["backgroundtest", "background test"],
    "LEAK TEST(COARSE)": ["leaktest(coarse)", "leak test coarse"],
    "LEAK TEST(FINE)": ["leaktest(fine)", "leak test fine", "fine leak test"],
    "TEST RESULT": ["testresult", "test result", "result"],
    "COARSE TEST START PRESSURE": ["coarsestartpressure", "coarse test start pressure"],
    "COARSE TEST END PRESSURE": ["coarseendpressure", "coarse test end pressure"],
    "COARSE TEST PRESSURE DROP": ["coarsepressuredrop", "coarse test pressure drop"],
    "PRESSURE TEST START PRESSURE": ["pressurestartpressure", "pressure test start pressure"],
    "PRESSURE TEST END PRESSURE": ["pressureendpressure", "pressure test end pressure"],
    "PRESSURE TEST PRESSURE DROP": ["pressurepressuredrop", "pressure test pressure drop"],
    "HE BACKGROUND": ["hebackground", "he background"],
    "ACTIVE MACHINE FACTOR": ["activemachinefactor", "active machine factor"],
    "HELIUM CONCENTERATION": ["heliumconcenteration", "helium concentration"],
    "HE COARSE FILL PRESSURE": ["hecoarsefillpressure", "he coarse fill pressure"],
    "HE COARSE TEST ELEPSED TIME": ["hecoarsetestelepsedtime", "he coarse test elepsed time"],
    "COARSE LEAK RATE(RAW VALUE)": ["coarseleakrate(rawvalue)", "coarse leak rate raw"],
    "COARSE LEAK RATE(CORRECTED)": ["coarseleakrate(corrected)", "coarse leak rate corrected"],
    "COARSE LEAK TEST VACUUM": ["coarseleaktestvacuum", "coarse leak test vacuum"],
    "HE FINE FILL PRESSURE": ["hefinefillpressure", "he fine fill pressure"],
    "HE FINE TEST ELEPSED TIME": ["hefinetestelepsedtime", "he fine test elepsed time"],
    "FINE LEAK RATE(RAW VALUE)": ["fineleakrate(rawvalue)", "fine leak rate raw"],
    "FINE LEAK RATE(CORRECTED)": ["fineleakrate(corrected)", "fine leak rate corrected"],
    "FINE LEAK TEST VACUUM": ["fineleaktestvacuum", "fine leak test vacuum"],
    "DRAWING NUMBER": ["drawingnumber", "drawing number"],
    "REFERENCE DOCUMENT": ["referencedocument", "reference document"],
    "CUSTOMER NAME": ["customername", "customer name"],
    "RECIPE NAME": ["recipename", "recipe name"],
    "COMPONENT NAME": ["componentname", "component name"],
    "SUPERVISOR": ["supervisor", "manager"]
}

editable_excel_columns = ["B", "D", "E", "F", "G", "I"]

def clone_row_format(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        src_cell = ws.cell(row=source_row, column=col)
        tgt_cell = ws.cell(row=target_row, column=col)
        if not isinstance(src_cell, MergedCell):
            tgt_cell.font = src_cell.font.copy()
            tgt_cell.border = src_cell.border.copy()
            tgt_cell.fill = src_cell.fill.copy()
            tgt_cell.alignment = src_cell.alignment.copy()
            tgt_cell.number_format = src_cell.number_format
            tgt_cell.protection = src_cell.protection.copy()
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row == source_row and merged_range.max_row == source_row:
            ws.merge_cells(
                start_row=target_row, start_column=merged_range.min_col,
                end_row=target_row, end_column=merged_range.max_col
            )

def parse_test_ids(raw_input):
    lines = raw_input.replace(',', '\n').split('\n')
    return [t.strip() for t in lines if t.strip()]

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        uploaded_file = request.files.get("file")
        test_ids_raw = request.form.get("test_ids", "")
        test_ids = parse_test_ids(test_ids_raw)
        action = request.form.get("action")

        if not uploaded_file or uploaded_file.filename == "":
            flash("❌ Please upload a valid Excel file.")
            return redirect(url_for("index"))

        if action == "generate" and not test_ids:
            flash("❌ Please enter at least one Identification Number.")
            return redirect(url_for("index"))

        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
                uploaded_file.save(temp_file.name)
                df = pd.read_excel(temp_file.name, dtype=str)

            normalized_cols = {normalize_key(col): col for col in df.columns}
            resolved_cols = {}
            for standard_name, variants in column_variants.items():
                for var in variants:
                    norm = normalize_key(var)
                    if norm in normalized_cols:
                        resolved_cols[standard_name] = normalized_cols[norm]
                        break

            if "IDENTIFICATION NUMBER" not in resolved_cols:
                flash("❌ No column matching 'Identification Number' found.")
                return redirect(url_for("index"))

            matched_col = resolved_cols["IDENTIFICATION NUMBER"]
            df[matched_col] = df[matched_col].astype(str).str.strip()

            if action == "add_all":
                matched_input_ids = df[matched_col].dropna().astype(str).tolist()
            else:
                matched_input_ids = [tid for tid in test_ids if tid in df[matched_col].values]

            template_path = os.path.join("static", "GIS BCT HELIUM.xlsx")
            output_path = os.path.join(tempfile.gettempdir(), "helium_output.xlsx")
            shutil.copy(template_path, output_path)

            wb = load_workbook(output_path)
            ws = wb.active

            # ✅ Insert image at K1
            logo_path = os.path.join("static", "logo.png")
            if os.path.exists(logo_path):
                img = ExcelImage(logo_path)
                img.width = 40
                img.height = 50
                ws.add_image(img, "L1")

            start_row = 5
            for i, ident_id in enumerate(matched_input_ids):
                target_row = start_row + i
                clone_row_format(ws, 5, target_row)
                row_data = df[df[matched_col] == ident_id].iloc[0]
                ws[f"A{target_row}"].value = i + 1
                for field, col_letter in cell_map.items():
                    input_col = resolved_cols.get(field)
                    cell = ws[f"{col_letter}{target_row}"]
                    if input_col and not isinstance(cell, MergedCell):
                        cell.value = row_data.get(input_col, "")
                        cell.protection = Protection(locked=True)
                for col in editable_excel_columns:
                    ws[f"{col}{target_row}"].protection = Protection(locked=False)

            ws.freeze_panes = "A5"
            ws.sheet_view.topLeftCell = "A1"
            ws.protection.sheet = True
            ws.protection.password = "locked123"
            wb.save(output_path)

            return send_file(output_path, as_attachment=True, download_name="helium_identification_export.xlsx")

        except Exception as e:
            flash(f"❌ Error: {str(e)}")
            return redirect(url_for("index"))

    return render_template("index.html")

if __name__ == "__main__":
    app.run(debug=True)


# final